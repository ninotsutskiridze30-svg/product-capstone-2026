#!/usr/bin/env python3
"""
Build the 6-month growth projection and 12-month financial model xlsx files.

Sources of truth for assumptions: ../04-gtm/financials/unit-economics.md
After editing assumptions there, re-run this script:

    python3 scripts/build-financial-models.py

Outputs:
- 04-gtm/growth-projection.xlsx       (3 scenarios × 6 months, with assumptions tab)
- 04-gtm/financials/twelve-month-model.xlsx  (12-month P&L with revenue and cost structure)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_GROWTH = ROOT / "04-gtm" / "growth-projection.xlsx"
OUT_TWELVE = ROOT / "04-gtm" / "financials" / "twelve-month-model.xlsx"

# ---------- Shared assumptions (must match ../04-gtm/financials/unit-economics.md) ----------
ASSUMPTIONS = {
    "Hourly rate (GEL)":                ("A1", 35,    "estimate", "Range observed across 4 tutor interviews (Apr 2026)"),
    "Lesson length (min)":              ("A2", 60,    "estimate", "Sprint 1 default; confirmed in interviews"),
    "Platform take rate":               ("A3", 0.15,  "strategic", "Positioned vs Preply (18-33%), iTalki (15%), Tutorful (22%)"),
    "Revenue / lesson (GEL)":           ("A4", 35*0.15, "derived",  "A1 × A3"),
    "Lessons / student / week":         ("A5", 2.0,   "benchmark", "Preply marketplace report 2024 — exam-prep avg 1.8-2.3"),
    "Active student lifespan (weeks)":  ("A6", 24,    "benchmark", "Georgian exam cycle Oct-Mar (~24 wks)"),
    "LiveKit USD / participant-min":    ("A7", 0.004, "measured",  "LiveKit Cloud pricing page May 2026"),
    "Lesson participants":              ("A8", 2,     "measured",  "1-on-1 lessons only in MVP"),
    "GEL per USD":                      ("FX", 2.65,  "measured",  "Bank of Georgia mid-rate May 2026"),
    "Variable infra GEL / lesson":      ("A9", 60*2*0.004*2.65, "derived", "A7 × A8 × A2 × FX"),
    "Storage GEL / lesson":             ("A10", 0.05, "estimate",  "5 MB/lesson × Supabase Storage pricing"),
}
ASSUMPTIONS["Total var cost / lesson (GEL)"] = ("A11", ASSUMPTIONS["Variable infra GEL / lesson"][1] + ASSUMPTIONS["Storage GEL / lesson"][1], "derived", "A9 + A10")
ASSUMPTIONS["Contribution / lesson (GEL)"] = ("A12", ASSUMPTIONS["Revenue / lesson (GEL)"][1] - ASSUMPTIONS["Total var cost / lesson (GEL)"][1], "derived", "A4 - A11")

# Channel mix at steady state (Month 6), expected scenario
CHANNEL_MIX = {"Tutor referral": 0.50, "Telegram content": 0.30, "IG/TikTok content": 0.20}
CHANNEL_CASH_CAC = {"Tutor referral": 1.32, "Telegram content": 0.0, "IG/TikTok content": 2.0}

# Fixed costs per month (GEL)
FIXED_COSTS = {
    "Vercel Hobby (free MVP) → Pro from M4 ($20/mo)": [0, 0, 0, 20*2.65, 20*2.65, 20*2.65, 20*2.65, 20*2.65, 20*2.65, 20*2.65, 20*2.65, 20*2.65],
    "Supabase Free → Pro from M4 ($25/mo)":           [0, 0, 0, 25*2.65, 25*2.65, 25*2.65, 25*2.65, 25*2.65, 25*2.65, 25*2.65, 25*2.65, 25*2.65],
    "Domain ($15/yr)":                                [15*2.65/12]*12,
    "PostHog (free tier sufficient through M12)":     [0]*12,
    "Founder opportunity cost (GEL 2000/mo)":         [2000]*12,
}

# ---------- Helpers ----------
THIN = Side(border_style="thin", color="d1d5db")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1e40af")
HEADER_FONT = Font(bold=True, color="ffffff")
TOTAL_FILL = PatternFill("solid", fgColor="dbeafe")
TOTAL_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="e0e7ff")
ASSUMP_FILL = PatternFill("solid", fgColor="fef3c7")

def autosize(ws, padding=2):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + padding, 12), 50)

def write_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

def write_assumptions_sheet(wb, name="Assumptions"):
    ws = wb.create_sheet(name)
    write_header(ws, 1, ["ID", "Input", "Value", "Class", "Source / Notes"])
    row = 2
    for label, (aid, val, cls, note) in ASSUMPTIONS.items():
        ws.cell(row=row, column=1, value=aid).border = BORDER
        ws.cell(row=row, column=2, value=label).border = BORDER
        c = ws.cell(row=row, column=3, value=val)
        c.border = BORDER
        if isinstance(val, float):
            c.number_format = "#,##0.0000" if abs(val) < 1 else "#,##0.00"
        cl = ws.cell(row=row, column=4, value=cls)
        cl.border = BORDER
        cl.fill = ASSUMP_FILL
        ws.cell(row=row, column=5, value=note).border = BORDER
        row += 1
    ws.cell(row=row+1, column=1, value="See ../financials/unit-economics.md for full prose and arithmetic.").font = Font(italic=True, color="6b7280")
    autosize(ws)
    return ws

# ---------- Sheet 1: 6-month projection × 3 scenarios ----------
SCENARIOS = {
    # (scenario_label,  starting_active_students_M1, monthly_growth_rate, tutor_seed_count)
    "Worst":    (5,   0.30, 4),   # 5 students M1, +30%/mo, 4 starter tutors
    "Expected": (10,  0.60, 8),   # 10 students M1, +60%/mo, 8 starter tutors
    "Best":     (20,  1.00, 15),  # 20 students M1, +100%/mo (low-base), 15 starter tutors
}

def build_growth_projection():
    wb = Workbook()
    wb.remove(wb.active)

    for label, (start, growth, tutor_seed) in SCENARIOS.items():
        ws = wb.create_sheet(label)
        ws.cell(row=1, column=1, value=f"6-Month Growth Projection — {label} scenario").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Start: {start} active students M1 · Monthly growth: {int(growth*100)}% · Starter tutors: {tutor_seed}").font = Font(italic=True, color="6b7280")

        headers = ["Metric", "M1", "M2", "M3", "M4", "M5", "M6", "Total / End"]
        write_header(ws, 4, headers)

        students = [round(start * ((1 + growth) ** i)) for i in range(6)]
        tutors   = [round(tutor_seed * ((1 + growth/2) ** i)) for i in range(6)]
        lessons  = [round(s * ASSUMPTIONS["Lessons / student / week"][1] * 4) for s in students]  # weeks per month ~4
        revenue  = [round(l * ASSUMPTIONS["Revenue / lesson (GEL)"][1], 2) for l in lessons]
        infra    = [round(l * ASSUMPTIONS["Total var cost / lesson (GEL)"][1], 2) for l in lessons]
        contrib  = [round(r - i, 2) for r, i in zip(revenue, infra)]
        new_signups = [students[0]] + [max(0, students[i] - students[i-1]) for i in range(1, 6)]
        # CAC: assume blended cash CAC from unit-economics
        blended_cac = sum(CHANNEL_MIX[c] * CHANNEL_CASH_CAC[c] for c in CHANNEL_MIX)
        cac_spend = [round(s * blended_cac, 2) for s in new_signups]

        def write_row(label, values, fmt=None, totalize=True, fill=None, bold=False):
            nonlocal row
            c = ws.cell(row=row, column=1, value=label)
            c.border = BORDER
            if fill: c.fill = fill
            if bold: c.font = Font(bold=True)
            for i, v in enumerate(values):
                cv = ws.cell(row=row, column=2+i, value=v)
                cv.border = BORDER
                if fmt: cv.number_format = fmt
                if fill: cv.fill = fill
                if bold: cv.font = Font(bold=True)
            if totalize:
                total = sum(values) if values and isinstance(values[0], (int, float)) else ""
                tv = ws.cell(row=row, column=8, value=total)
                tv.border = BORDER
                if fmt: tv.number_format = fmt
                if fill: tv.fill = fill
                tv.font = Font(bold=True)
            else:
                # "End" — last value
                ev = ws.cell(row=row, column=8, value=values[-1] if values else "")
                ev.border = BORDER
                if fmt: ev.number_format = fmt
                if fill: ev.fill = fill
                ev.font = Font(bold=True)
            row += 1

        row = 5
        write_row("Active students (end of month)", students, fmt="#,##0", totalize=False, fill=SECTION_FILL)
        write_row("Active tutors (end of month)",   tutors,   fmt="#,##0", totalize=False, fill=SECTION_FILL)
        write_row("New signups (students)",         new_signups, fmt="#,##0", totalize=True)
        write_row("Completed lessons",              lessons,  fmt="#,##0", totalize=True)
        write_row("Gross revenue (GEL)",            revenue,  fmt="#,##0.00", totalize=True)
        write_row("Variable infra cost (GEL)",      infra,    fmt="#,##0.00", totalize=True)
        write_row("Contribution margin (GEL)",      contrib,  fmt="#,##0.00", totalize=True, fill=TOTAL_FILL, bold=True)
        write_row("Acquisition spend (GEL, cash)",  cac_spend, fmt="#,##0.00", totalize=True)
        write_row("Contribution − CAC (GEL)",       [round(c - a, 2) for c, a in zip(contrib, cac_spend)],
                  fmt="#,##0.00", totalize=True, fill=TOTAL_FILL, bold=True)

        # Notes row
        ws.cell(row=row+1, column=1, value="Assumptions sourced from the Assumptions sheet. Scenario growth rates labelled as estimates.").font = Font(italic=True, color="6b7280")
        autosize(ws)

    # Comparison sheet
    cmp_ws = wb.create_sheet("Comparison", 0)
    cmp_ws.cell(row=1, column=1, value="6-Month Projection — scenario comparison at Month 6").font = Font(bold=True, size=14)
    headers = ["Metric", "Worst", "Expected", "Best"]
    write_header(cmp_ws, 3, headers)
    metrics = ["Active students", "Active tutors", "Monthly lessons", "Monthly revenue (GEL)", "Monthly contribution (GEL)"]
    rows_data = []
    for label, (start, growth, tutor_seed) in SCENARIOS.items():
        s6 = round(start * ((1 + growth) ** 5))
        t6 = round(tutor_seed * ((1 + growth/2) ** 5))
        l6 = round(s6 * ASSUMPTIONS["Lessons / student / week"][1] * 4)
        r6 = round(l6 * ASSUMPTIONS["Revenue / lesson (GEL)"][1], 2)
        c6 = round(r6 - l6 * ASSUMPTIONS["Total var cost / lesson (GEL)"][1], 2)
        rows_data.append((label, s6, t6, l6, r6, c6))
    by_label = {r[0]: r[1:] for r in rows_data}
    for i, m in enumerate(metrics):
        r = 4 + i
        cmp_ws.cell(row=r, column=1, value=m).border = BORDER
        for j, sc in enumerate(["Worst", "Expected", "Best"]):
            v = by_label[sc][i]
            cv = cmp_ws.cell(row=r, column=2+j, value=v)
            cv.border = BORDER
            cv.number_format = "#,##0" if i < 3 else "#,##0.00"
    autosize(cmp_ws)

    write_assumptions_sheet(wb, "Assumptions")
    OUT_GROWTH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_GROWTH)
    print(f"wrote {OUT_GROWTH.relative_to(ROOT)}")

# ---------- Sheet 2: 12-month financial model (Expected scenario) ----------
def build_twelve_month_model():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws.cell(row=1, column=1, value="12-Month Financial Model — Expected scenario").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="GEL · sourced assumptions on Assumptions tab · re-run scripts/build-financial-models.py to refresh").font = Font(italic=True, color="6b7280")

    headers = ["Metric"] + [f"M{i+1}" for i in range(12)] + ["Year 1 Total"]
    write_header(ws, 4, headers)

    start, growth, tutor_seed = SCENARIOS["Expected"]
    # Years 7-12: dampen growth as channels saturate
    months_growth = [growth] * 6 + [growth*0.6, growth*0.5, growth*0.4, growth*0.35, growth*0.3, growth*0.25]
    students = [start]
    for g in months_growth[1:]:
        students.append(round(students[-1] * (1 + g)))
    # tutors grow more slowly than students (each tutor serves multiple)
    tutors = [tutor_seed]
    for g in months_growth[1:]:
        tutors.append(round(tutors[-1] * (1 + g/2)))

    lessons = [round(s * ASSUMPTIONS["Lessons / student / week"][1] * 4) for s in students]
    revenue = [round(l * ASSUMPTIONS["Revenue / lesson (GEL)"][1], 2) for l in lessons]
    infra   = [round(l * ASSUMPTIONS["Total var cost / lesson (GEL)"][1], 2) for l in lessons]
    contrib = [round(r - i, 2) for r, i in zip(revenue, infra)]
    new_signups = [students[0]] + [max(0, students[i]-students[i-1]) for i in range(1, 12)]
    blended_cac = sum(CHANNEL_MIX[c] * CHANNEL_CASH_CAC[c] for c in CHANNEL_MIX)
    cac_spend = [round(s * blended_cac, 2) for s in new_signups]

    def write_row(label, values, fmt=None, fill=None, bold=False, total=True):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label); c.border = BORDER
        if fill: c.fill = fill
        if bold: c.font = Font(bold=True)
        for i, v in enumerate(values):
            cv = ws.cell(row=row, column=2+i, value=v); cv.border = BORDER
            if fmt: cv.number_format = fmt
            if fill: cv.fill = fill
            if bold: cv.font = Font(bold=True)
        tv = ws.cell(row=row, column=14, value=sum(values) if total else values[-1])
        tv.border = BORDER
        if fmt: tv.number_format = fmt
        if fill: tv.fill = fill
        tv.font = Font(bold=True)
        row += 1

    row = 5
    # Section: Revenue
    sc = ws.cell(row=row, column=1, value="REVENUE"); sc.font = Font(bold=True); sc.fill = SECTION_FILL; row += 1
    write_row("Active students (end of month)", students, fmt="#,##0", total=False, fill=None)
    write_row("Active tutors (end of month)",   tutors,   fmt="#,##0", total=False)
    write_row("Completed lessons",              lessons,  fmt="#,##0", total=True)
    write_row("Gross revenue (GEL)",            revenue,  fmt="#,##0.00", total=True, fill=TOTAL_FILL, bold=True)

    # Section: Variable costs
    sc = ws.cell(row=row, column=1, value="VARIABLE COSTS"); sc.font = Font(bold=True); sc.fill = SECTION_FILL; row += 1
    write_row("LiveKit + storage (GEL)",        infra,    fmt="#,##0.00", total=True)
    write_row("Acquisition spend (GEL, cash)",  cac_spend, fmt="#,##0.00", total=True)

    # Section: Fixed costs
    sc = ws.cell(row=row, column=1, value="FIXED COSTS"); sc.font = Font(bold=True); sc.fill = SECTION_FILL; row += 1
    total_fixed = [0.0]*12
    for label, vals in FIXED_COSTS.items():
        write_row(label, [round(v, 2) for v in vals], fmt="#,##0.00", total=True)
        for i, v in enumerate(vals): total_fixed[i] += v
    write_row("Total fixed costs (GEL)", [round(v, 2) for v in total_fixed], fmt="#,##0.00", total=True, fill=TOTAL_FILL, bold=True)

    # Bottom line
    sc = ws.cell(row=row, column=1, value="P&L"); sc.font = Font(bold=True); sc.fill = SECTION_FILL; row += 1
    total_costs = [round(infra[i] + cac_spend[i] + total_fixed[i], 2) for i in range(12)]
    net = [round(revenue[i] - total_costs[i], 2) for i in range(12)]
    write_row("Total costs (GEL)",   total_costs, fmt="#,##0.00", total=True)
    write_row("Net result (GEL)",    net,         fmt="#,##0.00", total=True, fill=TOTAL_FILL, bold=True)

    ws.cell(row=row+1, column=1, value="Net result is negative in early months by design — founder time is loaded at GEL 2000/mo opportunity cost. Cash break-even (excluding founder time) is reached when lesson contribution exceeds infra + paid CAC, see model. Re-run script with edited assumptions to refresh.").font = Font(italic=True, color="6b7280")
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=14)
    ws.cell(row=row+1, column=1).alignment = Alignment(wrap_text=True)

    autosize(ws)
    write_assumptions_sheet(wb, "Assumptions")

    OUT_TWELVE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_TWELVE)
    print(f"wrote {OUT_TWELVE.relative_to(ROOT)}")

if __name__ == "__main__":
    build_growth_projection()
    build_twelve_month_model()
